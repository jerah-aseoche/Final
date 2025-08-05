# backend/app/services/excel_filler.py
import io, os, re
from datetime import datetime
from typing import Dict, Any, Tuple

import pandas as pd
from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet

PLACEHOLDER_RE = re.compile(r"\{([^}]+)\}")

class ExcelTemplateFiller:
    def __init__(self, template_path: str, default_mapping: Dict[str, Any] | None = None):
        self.template_path = template_path
        self.default_mapping = default_mapping or {}

    def generate_from_filestorage(self, file_storage, mapping_json: str | None = None) -> Tuple[io.BytesIO, str]:
        mapping = self._merge_mapping(mapping_json)
        xl = self._read_uploaded_excel(file_storage)
        if len(xl) < 2:
            raise ValueError("Uploaded file must have at least 2 worksheets: details and grades.")

        sheet_names = list(xl.keys())
        df_details = xl[sheet_names[0]]
        df_grades  = xl[sheet_names[1]]

        if df_details.empty:
            raise ValueError("Details sheet has no rows.")

        wb, template_ws = self._load_template(self.template_path)

        used_titles = set()
        for idx, row in df_details.iterrows():
            row_dict = row.to_dict()
            name_key = next((k for k in mapping.keys() if k.upper() == "NAME"), "NAME")
            col_for_name = mapping.get(name_key, name_key)
            candidate_name = row_dict.get(col_for_name, f"Row {idx+1}")
            new_title = self._safe_sheet_title(str(candidate_name), used_titles)

            ws_copy = self._copy_template_sheet_with_fallback(wb, template_ws, new_title)

            matched_grades = df_grades[df_grades[col_for_name] == candidate_name]
            grade_row = matched_grades.iloc[0].to_dict() if not matched_grades.empty else {}
            combined_row = {**row_dict, **grade_row}
            self._replace_placeholders_in_worksheet(ws_copy, mapping, combined_row)

        wb.remove(template_ws)

        out = io.BytesIO()
        wb.save(out)
        out.seek(0)
        out_name = f"filled_multi_sheets_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        return out, out_name

    # ---- helpers ----
    def _merge_mapping(self, mapping_json: str | None) -> Dict[str, Any]:
        import json
        mapping = dict(self.default_mapping)
        if mapping_json and mapping_json.strip():
            mapping.update(json.loads(mapping_json))
        return mapping

    def _read_uploaded_excel(self, file_storage):
        xl = pd.read_excel(file_storage, sheet_name=None, dtype=str)
        return {k: v.fillna("") for k, v in xl.items()}

    def _load_template(self, template_path: str):
        if not os.path.exists(template_path):
            raise FileNotFoundError(f"Template not found on server: {template_path}")
        wb = load_workbook(template_path, data_only=True)
        if not wb.worksheets:
            raise RuntimeError("Template has no worksheets.")
        return wb, wb.worksheets[0]

    def _replace_placeholders_in_worksheet(self, ws: Worksheet, mapping: Dict[str, Any], rowdict: Dict[str, Any]):
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
            for cell in row:
                if isinstance(cell.value, str) and "{" in cell.value and "}" in cell.value:
                    cell.value = self._replace_placeholders_in_cell(cell.value, mapping, rowdict)

    def _replace_placeholders_in_cell(self, text: str, mapping: Dict[str, Any], rowdict: Dict[str, Any]) -> str:
        context = None
        up = text.upper()
        if "YEAR LAST ATTENDED" in up:
            if "ELEMENTARY" in up: context = "ELEMENTARY"
            elif "SECONDARY" in up: context = "SECONDARY"
            elif "TERTIARY" in up: context = "TERTIARY"

        def repl(m):
            key = m.group(1)
            mp = mapping.get(key, key)
            col = (mp.get(context) or mp.get("DEFAULT")) if isinstance(mp, dict) else mp
            val = rowdict.get(col, "")
            return "" if val is None else str(val)

        return PLACEHOLDER_RE.sub(repl, text)

    def _safe_sheet_title(self, s: str, used: set) -> str:
        title = (s or "").strip() or "Row"
        for ch in '[]:*?/\\':
            title = title.replace(ch, "-")
        title = title[:31] or "Row"
        base = title; i = 2
        while title in used:
            suffix = f" ({i})"
            title = (base[: 31 - len(suffix)] + suffix) if len(base) + len(suffix) > 31 else base + suffix
            i += 1
        used.add(title)
        return title

    def _copy_template_sheet_with_fallback(self, wb, template_ws, new_title: str):
        try:
            ws_copy = wb.copy_worksheet(template_ws)
            ws_copy.title = new_title
            return ws_copy
        except Exception:
            ws = wb.create_sheet(title=new_title)
            for rng in template_ws.merged_cells.ranges:
                ws.merge_cells(str(rng))
            for r in range(1, template_ws.max_row + 1):
                for c in range(1, template_ws.max_column + 1):
                    v = template_ws.cell(row=r, column=c).value
                    if v is not None:
                        ws.cell(row=r, column=c, value=v)
            return ws
